#!/usr/bin/env python3
"""
MSL Generator – India Priority Bucket Segments
===============================================
Reads the latest all_segments_*.csv from the outputs/ directory, extracts
distinct priority_bucket values, joins with India_Synthetic_SKU_Data.csv,
and writes one Excel workbook with one MSL sheet per priority bucket.

USAGE
-----
  python3 agents/msl_generator.py

OUTPUTS
-------
  MSL_Priority_Buckets_<timestamp>.xlsx
"""

import os
import glob
import sys
from collections import defaultdict
from datetime import datetime

try:
    import pandas as pd
except ImportError:
    sys.exit("pandas not installed. Run: pip install pandas --break-system-packages")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl --break-system-packages")


# ============================================================
# CONFIG
# ============================================================
_AGENTS_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(_AGENTS_DIR)          # one level up from agents/
OUTPUTS_DIR  = os.path.join(PROJECT_ROOT, "outputs")
# SKU data generated from Market_Master_File.csv outlets
SKU_CSV      = os.path.join(PROJECT_ROOT, "inbox", "India_Synthetic_SKU_Data.csv")
# Outlet master used to resolve IDs when the segmentation file lacks metadata
MARKET_MASTER_CSV = os.path.join(PROJECT_ROOT, "inbox", "Market_Master_File.csv")

# Top-N thresholds for cumulative business contribution summary
TOP_N_THRESHOLDS = [20, 35, 50]


# ============================================================
# COLOURS & STYLES
# ============================================================
C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "BDD7EE"
C_ORANGE_HDR = "F4B942"
C_YELLOW_HDR = "FFD966"
C_GRAY_XX    = "BFBFBF"
C_GREEN      = "92D050"
C_ORANGE     = "FFC000"
C_RED        = "FF0000"
C_WHITE      = "FFFFFF"
C_BLACK      = "000000"
C_LIGHT_GRAY = "F2F2F2"


def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)


def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def write_cell(ws, row, col, value=None, bold=False, fnt_color=C_BLACK,
               bg_color=None, num_fmt=None, h_align="left", v_align="center",
               wrap=False, bdr=None, size=10, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, color=fnt_color, size=size, italic=italic)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if bg_color:
        cell.fill = fill(bg_color)
    if num_fmt:
        cell.number_format = num_fmt
    if bdr:
        cell.border = bdr
    return cell


# ============================================================
# 1. FIND LATEST SEGMENTATION FILE
# ============================================================

def find_latest_segmentation_file(outputs_dir):
    """Return the path of the most recently modified all_segments_*.csv."""
    pattern = os.path.join(outputs_dir, "**", "all_segments_*.csv")
    candidates = glob.glob(pattern, recursive=True)
    if not candidates:
        sys.exit(f"No all_segments_*.csv found under {outputs_dir}")
    latest = max(candidates, key=os.path.getmtime)
    return latest


# ============================================================
# 2. LOAD & JOIN DATA
# ============================================================

def load_data(seg_path, sku_path):
    """
    Load segmentation file and India_Synthetic_SKU_Data.csv, join on outlet ID.

    India_Synthetic_SKU_Data.csv uses CUST_UNIQ_ID_VAL = OUTLET_UID_EDITED
    values from Market_Master_File.csv, so the join key is consistent across
    all three files.

    Returns a single DataFrame with columns:
      priority_bucket, BRND_NM, SUB_BRND_NM, FLVR_NM, PPG_DESC, NET_SALES,
      OUTLET_UID_EDITED, CUST_UNIQ_ID_VAL
    """
    seg = pd.read_csv(seg_path, low_memory=False)
    # India_Synthetic_SKU_Data.csv — generated from Market_Master_File.csv outlets
    sku = pd.read_csv(sku_path)

    if "OUTLET_UID_EDITED" not in seg.columns:
        sys.exit("Column 'OUTLET_UID_EDITED' not found in segmentation file.")
    if "priority_bucket" not in seg.columns:
        sys.exit("Column 'priority_bucket' not found in segmentation file.")
    if "CUST_UNIQ_ID_VAL" not in sku.columns:
        sys.exit(f"Column 'CUST_UNIQ_ID_VAL' not found in SKU file: {sku_path}")

    # Normalise IDs to strings to handle mixed int/string outlet IDs from
    # Market_Master_File.csv (e.g. 1477336 and S00865549)
    seg_slim = seg[["OUTLET_UID_EDITED", "priority_bucket"]].copy()
    seg_slim["OUTLET_UID_EDITED"] = seg_slim["OUTLET_UID_EDITED"].astype(str).str.strip()

    sku = sku.copy()
    sku["CUST_UNIQ_ID_VAL"] = sku["CUST_UNIQ_ID_VAL"].astype(str).str.strip()

    joined = sku.merge(
        seg_slim,
        left_on="CUST_UNIQ_ID_VAL",
        right_on="OUTLET_UID_EDITED",
        how="inner",
    )
    return joined


# ============================================================
# 3. COMPUTE PER-BUCKET METRICS
# ============================================================

def product_key(row):
    return f"{row['BRND_NM']} {row['SUB_BRND_NM']} {row['FLVR_NM']} ({row['PPG_DESC']})"


def build_metrics(bucket_df, all_df, bucket_stores):
    """
    Given data for one priority bucket:
      - bucket_df   : joined rows for this bucket
      - all_df      : all joined rows (for computing global share / index)
      - bucket_stores: set of all store IDs in this bucket (from segmentation)

    Returns:
      product_list, brand_data, ppg_data, ppg_order, brands_order, avg_sku
    """
    total_stores = len(bucket_stores)

    # ------------------------------------------------------------------
    # Product-level aggregation
    # ------------------------------------------------------------------
    prod_cols = ["BRND_NM", "SUB_BRND_NM", "FLVR_NM", "PPG_DESC"]

    # Sales per product in this bucket
    prod_sales_bucket = (
        bucket_df.groupby(prod_cols)["NET_SALES"].sum().reset_index()
    )
    bucket_total_sales = prod_sales_bucket["NET_SALES"].sum()

    # Sales per product across all buckets (for index)
    prod_sales_all = (
        all_df.groupby(prod_cols)["NET_SALES"].sum().reset_index()
    )
    all_total_sales = prod_sales_all["NET_SALES"].sum()

    # Store selling per product (how many stores in this bucket sell it)
    stores_per_product = (
        bucket_df.groupby(prod_cols)["CUST_UNIQ_ID_VAL"].nunique().reset_index()
        .rename(columns={"CUST_UNIQ_ID_VAL": "store_count"})
    )

    # Merge
    prod = prod_sales_bucket.merge(prod_sales_all[prod_cols + ["NET_SALES"]].rename(
        columns={"NET_SALES": "NET_SALES_ALL"}
    ), on=prod_cols, how="left")
    prod = prod.merge(stores_per_product, on=prod_cols, how="left")

    prod["cat_pct_mix"]      = prod["NET_SALES"] / bucket_total_sales if bucket_total_sales else 0
    prod["global_pct"]       = prod["NET_SALES_ALL"] / all_total_sales if all_total_sales else 0
    prod["index"]            = prod.apply(
        lambda r: (r["cat_pct_mix"] / r["global_pct"]) * 100 if r["global_pct"] else None,
        axis=1,
    )
    prod["pct_store_selling"] = prod["store_count"] / total_stores if total_stores else 0

    prod.sort_values("cat_pct_mix", ascending=False, inplace=True)
    prod.reset_index(drop=True, inplace=True)

    product_list = []
    for rank, (_, row) in enumerate(prod.iterrows(), 1):
        product_list.append({
            "product_name"       : product_key(row),
            "price_point"        : row["PPG_DESC"],
            "brand"              : row["BRND_NM"],
            "sub_cat"            : row["SUB_BRND_NM"],
            "flavour"            : row["FLVR_NM"],
            "mc_pct_nr"          : "",
            "msl_selection"      : "",
            "sku_type"           : "",
            "strategic_criterion": "",
            "msl_reasoning"      : "",
            "pct_store_selling"  : row["pct_store_selling"],
            "index"              : row["index"],
            "cat_pct_mix"        : row["cat_pct_mix"],
            "display_rank"       : rank,
        })

    # ------------------------------------------------------------------
    # Brand-level aggregation
    # ------------------------------------------------------------------
    brand_sales_bucket = bucket_df.groupby("BRND_NM")["NET_SALES"].sum()
    brand_sales_all    = all_df.groupby("BRND_NM")["NET_SALES"].sum()

    brand_data = {}
    for brand, bsales in brand_sales_bucket.items():
        b_pct_bucket = bsales / bucket_total_sales if bucket_total_sales else 0
        b_pct_all    = brand_sales_all.get(brand, 0) / all_total_sales if all_total_sales else 0
        brand_data[brand] = {
            "val_cont_pct": b_pct_bucket,
            "index"       : (b_pct_bucket / b_pct_all * 100) if b_pct_all else None,
        }

    brands_order = sorted(brand_data.keys(), key=lambda b: brand_data[b]["val_cont_pct"], reverse=True)

    # ------------------------------------------------------------------
    # PPG (price point) level aggregation
    # ------------------------------------------------------------------
    ppg_sales_bucket = bucket_df.groupby("PPG_DESC")["NET_SALES"].sum()
    ppg_sales_all    = all_df.groupby("PPG_DESC")["NET_SALES"].sum()

    ppg_data = {}
    for ppg, psales in ppg_sales_bucket.items():
        p_pct_bucket = psales / bucket_total_sales if bucket_total_sales else 0
        p_pct_all    = ppg_sales_all.get(ppg, 0) / all_total_sales if all_total_sales else 0
        ppg_data[ppg] = {
            "val_cont_pct": p_pct_bucket,
            "index"       : (p_pct_bucket / p_pct_all * 100) if p_pct_all else None,
        }

    ppg_order = sorted(ppg_data.keys(), key=lambda p: ppg_sales_bucket.get(p, 0), reverse=True)

    # Average SKUs per store
    avg_sku = round(bucket_df.groupby("CUST_UNIQ_ID_VAL")["NET_SALES"].count().mean(), 1)

    return product_list, brand_data, ppg_data, ppg_order, brands_order, avg_sku


def build_sku_matrix(product_list, brands_order, ppg_order):
    """Returns { brand: { ppg: count } }  and  { brand: total_count }."""
    matrix = defaultdict(lambda: defaultdict(int))
    for p in product_list:
        matrix[p["brand"]][p["price_point"]] += 1
    totals = {b: sum(matrix[b].values()) for b in brands_order}
    return matrix, totals


def cumulative_business(product_list, thresholds):
    result = {}
    running = 0.0
    for i, p in enumerate(product_list, 1):
        running += p["cat_pct_mix"] or 0
        if i in thresholds:
            result[i] = running
    return result


# ============================================================
# 4. LLM-DRIVEN MSL SELECTION
# ============================================================

_BUCKET_PROFILES = {
    "A": ("High-revenue, full-range stores",
          "Broadest MSL 15–20 SKUs. All Hero SKUs + full Strategic SKU set across all 8 criteria."),
    "B": ("Good-revenue, developing range",
          "Core MSL 10–15 SKUs. All Hero SKUs + Strategic SKUs driven by Execution, Profitability, Incrementality."),
    "C": ("Mid-revenue, selective range",
          "Focused MSL 8–12 SKUs. All Hero SKUs + Strategic SKUs limited to Consumer/Shopper Relevance, Trade Relevance."),
    "D": ("Low-revenue, basic range",
          "Minimal MSL 5–8 SKUs. Hero SKUs only (highest-penetration ₹5/₹10 price points). "
          "Strategic SKUs only if Innovation criterion applies (NPD/Big Bets)."),
}

_STRATEGIC_CRITERIA = [
    "Execution", "Consumer/Shopper Relevance", "Profitability",
    "Trade Relevance", "Incrementality", "Market Relevance", "Strategy", "Innovation",
]


def _fallback_msl_selection(product_list: list) -> None:
    """Data-driven fallback: Hero = top SKUs reaching ~50% cumulative mix."""
    running = 0.0
    for p in product_list:
        running += (p.get("cat_pct_mix") or 0)
        if running <= 0.50:
            p["sku_type"]            = "Hero"
            p["strategic_criterion"] = ""
            p["msl_reasoning"]       = f"Top revenue contributor; cumulative mix {running:.1%}"
            p["msl_selection"]       = "✓"
        else:
            p["sku_type"]            = ""
            p["strategic_criterion"] = ""
            p["msl_reasoning"]       = ""
            p["msl_selection"]       = ""


def llm_msl_selection(bucket_name: str, product_list: list,
                      api_key: str, model: str,
                      cluster_info: str = "") -> list:
    """
    Makes an LLM call to classify each SKU as Hero / Strategic / not-in-MSL,
    grounded in the MSL context rules and the product-level data.
    Modifies product_list in-place and returns it.
    Falls back to _fallback_msl_selection() on any error.
    """
    import sys as _sys
    _sys.path.insert(0, PROJECT_ROOT)
    try:
        from agents.llm_client import call_llm
        from agents.context_loader import ContextLoader
    except ImportError:
        print("[MSL] llm_client/context_loader not importable — using data-driven fallback")
        _fallback_msl_selection(product_list)
        return product_list

    base_bucket = bucket_name.rstrip("+").upper()[:1]
    profile_label, profile_desc = _BUCKET_PROFILES.get(
        base_bucket, (f"Priority Bucket {bucket_name}", "Standard MSL rules apply."))

    # Build product table (top 60 SKUs to stay within token budget)
    header = "RANK | PRODUCT | PPG | BRAND | % MIX | CUM MIX | % STORES | INDEX"
    rows = [header]
    running = 0.0
    for p in product_list[:60]:
        running += (p.get("cat_pct_mix") or 0)
        idx = p.get("index")
        rows.append(
            f"{p['display_rank']} | {p['product_name']} | {p['price_point']} | "
            f"{p['brand']} | {p['cat_pct_mix']:.1%} | {running:.1%} | "
            f"{p['pct_store_selling']:.1%} | {round(idx) if idx else 'N/A'}"
        )
    product_table = "\n".join(rows)

    cluster_block = f"\nCLUSTER CONTEXT:\n{cluster_info}\n" if cluster_info else ""

    prompt = f"""You are selecting the Must Stock List (MSL) for Priority Bucket {bucket_name}.

BUCKET PROFILE: {profile_label}
{profile_desc}
{cluster_block}
PRODUCT LIST (ranked by revenue mix contribution to this bucket):
{product_table}

TASK:
1. Hero SKUs — mark SKUs whose CUMULATIVE MIX reaches ~50%. These are non-negotiable for every store.
2. Strategic SKUs — beyond the Hero set, identify a small number of SKUs that qualify under at least one of these criteria (only criteria allowed for this bucket tier apply):
   {', '.join(_STRATEGIC_CRITERIA)}
   Strategic SKUs should target distribution uplift to gain revenue contribution or reflect cluster-specific shopper needs. Provide one short reasoning sentence grounded in the data (index, store penetration, PPG gap).
3. Not-in-MSL — leave all remaining SKUs blank.

MSL size guideline: A=15–20 SKUs, B=10–15 SKUs, C=8–12 SKUs, D=5–8 SKUs.

OUTPUT FORMAT — respond ONLY with a pipe-delimited table, one row per SKU, starting directly with the header line. No preamble, no text outside the table.

RANK | SKU_TYPE | CRITERION | REASONING

Rules:
- SKU_TYPE must be exactly: Hero, Strategic, or blank
- CRITERION must be one of the 8 criteria above, or blank (blank for Hero and not-in-MSL)
- REASONING: 1 short sentence for Hero and Strategic; blank for not-in-MSL
- Include ALL {min(len(product_list), 60)} product ranks in the table (even if blank)
"""

    ctx = ContextLoader()
    system_prompt = ctx.build("msl")

    try:
        response = call_llm(prompt, api_key, model, max_tokens=3000,
                            system_prompt=system_prompt)
        rank_map: dict = {}
        for line in response.strip().split("\n"):
            parts = [p.strip() for p in line.split("|")]
            if len(parts) < 2:
                continue
            try:
                rank = int(parts[0])
            except ValueError:
                continue  # header or blank
            rank_map[rank] = {
                "sku_type"           : parts[1] if len(parts) > 1 else "",
                "strategic_criterion": parts[2] if len(parts) > 2 else "",
                "msl_reasoning"      : parts[3] if len(parts) > 3 else "",
            }

        if not rank_map:
            raise ValueError("LLM returned no parseable rows")

        for p in product_list:
            info = rank_map.get(p["display_rank"], {})
            p["sku_type"]            = info.get("sku_type", "")
            p["strategic_criterion"] = info.get("strategic_criterion", "")
            p["msl_reasoning"]       = info.get("msl_reasoning", "")
            p["msl_selection"]       = "✓" if p["sku_type"] in ("Hero", "Strategic") else ""

    except Exception as e:
        print(f"[MSL] LLM selection failed for bucket {bucket_name} ({e}) — using data-driven fallback")
        _fallback_msl_selection(product_list)

    return product_list


# ============================================================
# 5. WRITE ONE EXCEL SHEET PER BUCKET
# ============================================================

def write_bucket_sheet(wb, bucket_name, product_list, brand_data, ppg_data,
                       ppg_order, brands_order, avg_sku):
    ws = wb.create_sheet(title=f"Bucket {bucket_name}"[:31])

    n_pp = len(ppg_order)

    # Column layout:
    #  Col 1: spacer | 2: index/label | 3: val_cont% | 4: brand/ppg label
    #  Cols 5..5+n_pp-1: ppg columns (left matrix)
    #  Col 5+n_pp: TOTAL (left matrix)
    #  Col 5+n_pp+1: spacer
    #  Cols 5+n_pp+2..5+2*n_pp+1: ppg columns (right MSL matrix)

    msl_offset = 5 + n_pp + 2   # first column of right MSL matrix

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 4
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 12
    ws.column_dimensions[get_column_letter(4)].width = 18
    for i in range(n_pp):
        ws.column_dimensions[get_column_letter(5 + i)].width = 13
    ws.column_dimensions[get_column_letter(5 + n_pp)].width = 8
    ws.column_dimensions[get_column_letter(5 + n_pp + 1)].width = 4
    for i in range(n_pp):
        ws.column_dimensions[get_column_letter(msl_offset + i)].width = 13

    def pp_col(i):  return 5 + i
    def msl_col(i): return msl_offset + i

    R = 1

    # ── Rows R1-R3: Bucket / Country / AvgSKU ────────────────────────────────
    for lr, label, val in [
        (R,   "Segment",  f"Priority Bucket {bucket_name}"),
        (R+1, "Country",  "India"),
        (R+2, "Avg SKU",  avg_sku),
    ]:
        write_cell(ws, lr, 2, label, bold=True, bg_color=C_MID_BLUE,
                   fnt_color=C_WHITE, h_align="center", bdr=border())
        write_cell(ws, lr, 3, val, bold=True, bg_color=C_LIGHT_BLUE,
                   h_align="center", bdr=border())

    # ── Row R+3: PPG Index values ─────────────────────────────────────────────
    write_cell(ws, R+3, 2, "Index", bold=True, bg_color=C_MID_BLUE,
               fnt_color=C_WHITE, h_align="center", bdr=border())
    for i, ppg in enumerate(ppg_order):
        val = (ppg_data.get(ppg) or {}).get("index")
        write_cell(ws, R+3, pp_col(i), round(val, 2) if val else "",
                   num_fmt="0.00", h_align="center", bdr=border(), bg_color=C_LIGHT_BLUE)
        write_cell(ws, R+3, msl_col(i), round(val, 2) if val else "",
                   num_fmt="0.00", h_align="center", bdr=border(), bg_color=C_LIGHT_BLUE)

    # ── Row R+4: Val cont% ────────────────────────────────────────────────────
    write_cell(ws, R+4, 3, "Val cont%", bg_color=C_LIGHT_BLUE,
               h_align="center", bdr=border())
    for i, ppg in enumerate(ppg_order):
        val = (ppg_data.get(ppg) or {}).get("val_cont_pct")
        write_cell(ws, R+4, pp_col(i), val,
                   num_fmt="0.00%", h_align="center", bdr=border(), bg_color=C_LIGHT_BLUE)
        write_cell(ws, R+4, msl_col(i), val,
                   num_fmt="0.00%", h_align="center", bdr=border(), bg_color=C_LIGHT_BLUE)

    # ── Row R+5: Column headers ───────────────────────────────────────────────
    write_cell(ws, R+5, 4, "PPG / Brand", bold=True, bg_color=C_ORANGE_HDR,
               h_align="center", wrap=True, bdr=border())
    write_cell(ws, R+5, msl_offset, "PPG / Brand", bold=True, bg_color=C_YELLOW_HDR,
               h_align="center", wrap=True, bdr=border())
    write_cell(ws, R+5, 2, "", bg_color=C_ORANGE_HDR, bdr=border())
    write_cell(ws, R+5, 3, "", bg_color=C_ORANGE_HDR, bdr=border())

    for i, ppg in enumerate(ppg_order):
        write_cell(ws, R+5, pp_col(i), ppg, bold=True, bg_color=C_ORANGE_HDR,
                   h_align="center", bdr=border())
        write_cell(ws, R+5, msl_col(i), ppg, bold=True, bg_color=C_YELLOW_HDR,
                   h_align="center", bdr=border())

    write_cell(ws, R+5, pp_col(n_pp), "TOTAL", bold=True, bg_color=C_ORANGE_HDR,
               h_align="center", bdr=border())

    # ── Brand rows ────────────────────────────────────────────────────────────
    matrix, matrix_totals = build_sku_matrix(product_list, brands_order, ppg_order)

    for b_idx, brand in enumerate(brands_order):
        row_r = R + 6 + b_idx
        bd      = brand_data.get(brand, {})
        idx_val = bd.get("index")
        vc_val  = bd.get("val_cont_pct")

        write_cell(ws, row_r, 2, round(idx_val, 2) if idx_val else "",
                   num_fmt="0.00", h_align="center", bdr=border())
        write_cell(ws, row_r, 3, vc_val, num_fmt="0.00%", h_align="center", bdr=border())
        write_cell(ws, row_r, 4, brand, bold=True, h_align="left", bdr=border())

        row_total = 0
        for i, ppg in enumerate(ppg_order):
            count = matrix[brand].get(ppg, 0)
            if count:
                write_cell(ws, row_r, pp_col(i), count, h_align="center", bdr=border())
                write_cell(ws, row_r, msl_col(i), "x", h_align="center",
                           bdr=border(), bg_color=C_YELLOW_HDR)
                row_total += count
            else:
                write_cell(ws, row_r, pp_col(i), "xx", h_align="center",
                           bdr=border(), bg_color=C_GRAY_XX)
                write_cell(ws, row_r, msl_col(i), "xx", h_align="center",
                           bdr=border(), bg_color=C_GRAY_XX)

        write_cell(ws, row_r, pp_col(n_pp), row_total, bold=True,
                   h_align="center", bdr=border())

    # ── 3 spare brand rows ────────────────────────────────────────────────────
    for spare in range(3):
        row_r = R + 6 + len(brands_order) + spare
        for col in [2, 3, 4] + list(range(pp_col(0), pp_col(n_pp) + 1)):
            write_cell(ws, row_r, col, "", bdr=border())
        for i in range(n_pp):
            write_cell(ws, row_r, msl_col(i), "xx", h_align="center",
                       bdr=border(), bg_color=C_GRAY_XX)

    # ── TOTAL row ─────────────────────────────────────────────────────────────
    total_row = R + 6 + len(brands_order) + 3
    grand_total = sum(matrix_totals.values())
    write_cell(ws, total_row, pp_col(n_pp), grand_total, bold=True,
               h_align="center", bg_color=C_LIGHT_BLUE, bdr=border())

    # ── Summary / Legend block ─────────────────────────────────────────────────
    SUM_R = total_row + 3

    write_cell(ws, SUM_R,   msl_offset, "≤ 55% : Red",     bg_color=C_RED,
               fnt_color=C_WHITE, h_align="center", bdr=border())
    write_cell(ws, SUM_R+1, msl_offset, "56–69% : Orange",  bg_color=C_ORANGE,
               h_align="center", bdr=border())
    write_cell(ws, SUM_R+2, msl_offset, "≥ 70% : Green",    bg_color=C_GREEN,
               h_align="center", bdr=border())

    write_cell(ws, SUM_R - 1, 6, "NEW MSL", bold=True, bg_color=C_MID_BLUE,
               fnt_color=C_WHITE, h_align="center")
    write_cell(ws, SUM_R,     6, "No. of SKU", bold=True, h_align="center")
    write_cell(ws, SUM_R,     7, "Cont %",     bold=True, h_align="center")
    write_cell(ws, SUM_R - 1, 9, "Top SKUs Ranking",    bold=True, h_align="center")
    write_cell(ws, SUM_R,     9, "Rank Threshold",      bold=True, h_align="center")
    write_cell(ws, SUM_R,    10, "Bus. Contribution",   bold=True, h_align="center", wrap=True)

    cum_biz = cumulative_business(product_list, TOP_N_THRESHOLDS)
    for row_offset, n in enumerate(TOP_N_THRESHOLDS, 1):
        cum = cum_biz.get(n, 0)
        write_cell(ws, SUM_R + row_offset, 9,  n,   h_align="center")
        write_cell(ws, SUM_R + row_offset, 10, cum, num_fmt="0.00%", h_align="center")

    # ── Product list table ────────────────────────────────────────────────────
    PROD_R = SUM_R + len(TOP_N_THRESHOLDS) + 4

    # Colours for MSL type rows
    C_HERO_BG  = "E2EFDA"   # light green  — Hero SKU rows
    C_STRAT_BG = "DDEEFF"   # light blue   — Strategic SKU rows

    prod_headers = [
        "PRODUCT DESCRIPTION",     # col 1
        "PPG",                     # col 2
        "BRAND",                   # col 3
        "SUB-BRAND",               # col 4
        "FLAVOUR",                 # col 5
        "MC %NR",                  # col 6
        "MSL SELECTION",           # col 7
        "SKU TYPE",                # col 8
        "STRATEGIC CRITERION",     # col 9
        "REASONING",               # col 10
        "% No of Store Selling",   # col 11
        "RANKING",                 # col 12
        "INDEX",                   # col 13
        "% MIX OF BUCKET (SHARE)", # col 14
    ]

    prod_col_widths = [45, 13, 14, 14, 30, 10, 10, 14, 24, 42, 22, 10, 10, 24]
    for ci, w in enumerate(prod_col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ci, hdr in enumerate(prod_headers, 1):
        write_cell(ws, PROD_R, ci, hdr,
                   bold=True, fnt_color=C_WHITE, bg_color=C_DARK_BLUE,
                   h_align="center", v_align="center", wrap=True,
                   bdr=border(), size=9)
    ws.row_dimensions[PROD_R].height = 30

    ws.freeze_panes = ws.cell(row=PROD_R + 1, column=2)

    for offset, p in enumerate(product_list):
        dr = PROD_R + 1 + offset
        bg = C_LIGHT_GRAY if offset % 2 == 1 else C_WHITE

        store_pct  = p.get("pct_store_selling")
        idx        = p.get("index")
        sku_type   = p.get("sku_type", "")
        criterion  = p.get("strategic_criterion", "")
        reasoning  = p.get("msl_reasoning", "")

        # Row highlight based on MSL type
        type_bg = C_HERO_BG if sku_type == "Hero" else C_STRAT_BG if sku_type == "Strategic" else bg

        row_vals = [
            (p["product_name"],                          None,      type_bg),  # col 1
            (p["price_point"],                           None,      type_bg),  # col 2
            (p["brand"],                                 None,      type_bg),  # col 3
            (p["sub_cat"],                               None,      type_bg),  # col 4
            (p["flavour"],                               None,      type_bg),  # col 5
            (p["mc_pct_nr"],                             None,      type_bg),  # col 6
            (p["msl_selection"],                         None,      type_bg),  # col 7
            (sku_type,                                   None,      type_bg),  # col 8
            (criterion,                                  None,      type_bg),  # col 9
            (reasoning,                                  None,      type_bg),  # col 10
            (store_pct,                                  "0.00%",   None),     # col 11
            (p["display_rank"],                          "0",       type_bg),  # col 12
            (round(idx, 2) if idx is not None else "",   "0.00",    type_bg),  # col 13
            (p["cat_pct_mix"],                           "0.00%",   type_bg),  # col 14
        ]

        for ci, (val, fmt, cell_bg) in enumerate(row_vals, 1):
            c = write_cell(ws, dr, ci, val, num_fmt=fmt,
                           bg_color=cell_bg, bdr=border(), size=9)
            if ci == 11:  # % No of Store Selling — traffic-light colouring
                if isinstance(store_pct, float):
                    if store_pct <= 0.55:
                        c.fill = fill(C_RED)
                    elif store_pct <= 0.69:
                        c.fill = fill(C_ORANGE)
                    else:
                        c.fill = fill(C_GREEN)
                else:
                    c.fill = fill(C_LIGHT_GRAY)

    ws.auto_filter.ref = (
        f"A{PROD_R}:{get_column_letter(len(prod_headers))}{PROD_R + len(product_list)}"
    )
    ws.print_title_rows = f"{PROD_R}:{PROD_R}"

    print(f"  [Bucket {bucket_name}] {len(product_list)} products | avg SKU/store: {avg_sku}")


# ============================================================
# 5. PIPELINE ENTRY POINT (called from pipeline.py)
# ============================================================

def run_msl_from_df(df_segments, sku_path: str, output_dir: str,
                    api_key: str = None, model: str = None,
                    labels: dict = None) -> str:
    """
    Pipeline-integrated MSL generation.

    Takes the in-memory segments DataFrame (must already carry OUTLET_UID_EDITED
    and priority_bucket from the prioritization step), joins with the SKU data
    file, and writes one MSL sheet per priority bucket to output_dir.

    api_key / model: when provided, an LLM call marks Hero vs Strategic SKUs.
    labels: segment labels dict from run_segmentation() — used as cluster context
            for the LLM reasoning prompt.

    Returns the saved workbook path, or "" if generation is skipped.
    """
    import logging
    log = logging.getLogger("perfect_store.msl")

    if not os.path.exists(sku_path):
        log.warning(f"[MSL] SKU file not found: {sku_path} — skipping MSL generation")
        return ""

    required = {"OUTLET_UID_EDITED", "priority_bucket"}
    missing = required - set(df_segments.columns)
    if missing:
        log.warning(f"[MSL] Missing columns in segments DataFrame: {missing} — skipping")
        return ""

    sku = pd.read_csv(sku_path)

    seg_slim = df_segments[["OUTLET_UID_EDITED", "priority_bucket"]].copy()
    if "segment_label" in df_segments.columns:
        seg_slim = df_segments[["OUTLET_UID_EDITED", "priority_bucket", "segment_label"]].copy()
    seg_slim["OUTLET_UID_EDITED"] = seg_slim["OUTLET_UID_EDITED"].astype(str).str.strip()
    sku["CUST_UNIQ_ID_VAL"] = sku["CUST_UNIQ_ID_VAL"].astype(str).str.strip()

    joined = sku.merge(
        seg_slim,
        left_on="CUST_UNIQ_ID_VAL",
        right_on="OUTLET_UID_EDITED",
        how="inner",
    )

    if joined.empty:
        log.warning("[MSL] No matching outlets between SKU data and segments — skipping")
        return ""

    buckets = sorted(joined["priority_bucket"].dropna().unique())
    log.info(f"[MSL] Generating sheets for buckets: {buckets}")

    bucket_store_sets = {
        b: set(seg_slim.loc[seg_slim["priority_bucket"] == b, "OUTLET_UID_EDITED"])
        for b in buckets
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for bucket in buckets:
        bucket_df = joined[joined["priority_bucket"] == bucket].copy()
        product_list, brand_data, ppg_data, ppg_order, brands_order, avg_sku = build_metrics(
            bucket_df, joined, bucket_store_sets[bucket]
        )

        # Build cluster context string for this bucket
        cluster_info = ""
        if "segment_label" in seg_slim.columns:
            top_segs = (
                seg_slim.loc[seg_slim["priority_bucket"] == bucket, "segment_label"]
                .value_counts().head(3)
            )
            cluster_info = "\n".join(
                f"- {seg} ({cnt} stores)" for seg, cnt in top_segs.items()
            )
        elif labels:
            cluster_lines = []
            for cid, info in labels.items():
                lbl = info.get("label", f"Cluster {cid}")
                identity = (info.get("identity_card") or "")[:120].replace("\n", " ")
                cluster_lines.append(f"- {lbl}: {identity}")
            cluster_info = "\n".join(cluster_lines[:6])

        if api_key:
            llm_model = model or "claude-opus-4-6"
            log.info(f"[MSL] LLM MSL selection for bucket {bucket} ...")
            llm_msl_selection(bucket, product_list, api_key, llm_model,
                              cluster_info=cluster_info)
        else:
            _fallback_msl_selection(product_list)

        write_bucket_sheet(wb, bucket, product_list, brand_data, ppg_data,
                           ppg_order, brands_order, avg_sku)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"MSL_Priority_Buckets_{timestamp}.xlsx")
    wb.save(output_path)
    log.info(f"[MSL] Workbook saved: {output_path}")
    return output_path


# ============================================================
# 6. STANDALONE MAIN (python agents/msl_generator.py)
# ============================================================

def main():
    # -- Load API key (optional — falls back to data-driven selection) ----------
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(PROJECT_ROOT, ".env"))
    except ImportError:
        pass
    llm_backend = os.getenv("LLM_BACKEND", "anthropic").strip().lower()
    if llm_backend == "azure":
        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        model   = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
    elif llm_backend == "gemini":
        api_key = os.getenv("GEMINI_API_KEY")
        model   = os.getenv("GEMINI_MODEL", "gemini-1.5-pro")
    elif llm_backend in ("local", "vertex"):
        api_key = None
        model   = os.getenv("LOCAL_LLM_MODEL") or os.getenv("VERTEX_MODEL", "gemini-1.5-pro")
    else:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        model   = os.getenv("CLAUDE_MODEL", "claude-opus-4-6")

    # -- Find latest segmentation file -----------------------------------------
    seg_path = find_latest_segmentation_file(OUTPUTS_DIR)
    print(f"Segmentation file : {seg_path}")
    print(f"SKU data          : {SKU_CSV}")
    print(f"LLM backend       : {llm_backend} | model: {model} | key: {'set' if api_key else 'not set'}")

    if not os.path.exists(SKU_CSV):
        sys.exit(f"SKU file not found: {SKU_CSV}")

    # -- Load & join -----------------------------------------------------------
    joined = load_data(seg_path, SKU_CSV)
    print(f"Joined rows       : {len(joined):,}")

    # -- Determine buckets (sorted A, B, C, D ...) ----------------------------
    buckets = sorted(joined["priority_bucket"].dropna().unique())
    print(f"Priority buckets  : {buckets}")

    # Store sets per bucket (from full segmentation, not just joined rows)
    seg_full = pd.read_csv(seg_path, low_memory=False)
    seg_full["OUTLET_UID_EDITED"] = seg_full["OUTLET_UID_EDITED"].astype(str).str.strip()
    bucket_store_sets = {
        b: set(seg_full.loc[seg_full["priority_bucket"] == b, "OUTLET_UID_EDITED"])
        for b in buckets
    }

    # Cluster context (segment_label from segmentation CSV if present)
    has_labels = "segment_label" in seg_full.columns

    # -- Build workbook --------------------------------------------------------
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for bucket in buckets:
        bucket_df = joined[joined["priority_bucket"] == bucket].copy()
        product_list, brand_data, ppg_data, ppg_order, brands_order, avg_sku = build_metrics(
            bucket_df, joined, bucket_store_sets[bucket]
        )

        cluster_info = ""
        if has_labels:
            top_segs = (
                seg_full.loc[seg_full["priority_bucket"] == bucket, "segment_label"]
                .value_counts().head(3)
            )
            cluster_info = "\n".join(
                f"- {seg} ({cnt} stores)" for seg, cnt in top_segs.items()
            )

        if api_key:
            print(f"  [Bucket {bucket}] Running LLM MSL selection...")
            llm_msl_selection(bucket, product_list, api_key, model,
                              cluster_info=cluster_info)
        else:
            print(f"  [Bucket {bucket}] No API key — using data-driven fallback")
            _fallback_msl_selection(product_list)

        write_bucket_sheet(
            wb, bucket, product_list, brand_data, ppg_data,
            ppg_order, brands_order, avg_sku,
        )

    # -- Save ------------------------------------------------------------------
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUTS_DIR, f"MSL_Priority_Buckets_{timestamp}.xlsx")
    wb.save(output_path)
    print(f"\n[OK] Saved: {output_path}")


if __name__ == "__main__":
    main()
